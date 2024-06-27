import datetime
import os
import sys
from PyQt6 import QtWidgets
from PyQt6.QtCore import QThread, Qt, pyqtSignal
from openpyxl.reader.excel import load_workbook


def read_excel_template(throw_exception, template_path="flat"):
    template_path = f"_internal/input/template_{template_path}.xlsm"
    if os.path.exists(template_path):
        book = load_workbook(template_path, keep_vba=True)
        return book
    else:
        throw_exception.emit("Повреждена файловая система! Обратитесь в поддержку")


def fill_table_pyqt(table, header, data, currency, data_start=0, data_finish=100, price_index=8):
    table.setColumnCount(len(header))
    data = data[data_start:data_finish]
    table.setRowCount(len(data))
    table.setHorizontalHeaderLabels(header)

    fill_table_data_pyqt(table, data)
    table.resizeColumnsToContents()
    if currency == 'uzs':
        table.setColumnHidden(price_index, False)
        table.setColumnHidden(price_index + 1, False)
        table.setColumnHidden(price_index + 2, True)
        table.setColumnHidden(price_index + 3, True)
    else:
        table.setColumnHidden(price_index + 3, False)
        table.setColumnHidden(price_index + 2, False)
        table.setColumnHidden(price_index + 1, True)
        table.setColumnHidden(price_index, True)


def fill_table_data_pyqt(table, data):
    for one_row in data:
        _one_row = one_row.prepare_to_list()
        for i in range(len(_one_row)):
            if _one_row[i] is None:
                item = QtWidgets.QTableWidgetItem("-")
            else:
                val = str(_one_row[i]) if len(str(_one_row[i])) < 50 else (str(_one_row[i])[:50] + "...")
                item = QtWidgets.QTableWidgetItem(val)
            # item.setFlags(Qt.ItemFlag.ItemIsEditable)
            table.setItem(data.index(one_row), i, item)
            item.setFlags(Qt.ItemFlag.ItemIsUserCheckable)


class Exporter(QThread):
    throw_exception = pyqtSignal(str)
    throw_info = pyqtSignal(str)
    block_closing = pyqtSignal(bool)

    def __init__(self, name, type_real_estate, path='output/', results=None):
        super().__init__()
        self.path = path
        self.name = ''
        self.results = results
        self.type_real_estate = type_real_estate

    def run(self):
        try:
            self.block_closing.emit(True)

            self.name = f"{datetime.datetime.now().strftime('%d-%m-%Y-%H-%M-%S')}_{self.type_real_estate}"
            book = read_excel_template(self.throw_exception, self.type_real_estate)
            sheet = book[book.sheetnames[0]]
            sheet.title = f"{datetime.datetime.now().strftime('%d.%m.%y_%H.%M')}"

            if len(self.results) == 0:
                self.throw_info.emit(f"Не найдены данные по запросу для {self.name}")
                return
            for i in range(0, len(self.results)):
                sheet.append(self.results[i].prepare_to_list())
            self.path += f'{self.name}.xlsm'
            try:
                if os.path.exists(self.path):
                    os.remove(self.path)
            except Exception as e:
                print(e)
            book.save(self.path)
            self.block_closing.emit(False)
        except Exception as e:
            print("Export", self.name, e)
