import math
import os
import random
import re

import time
# from threading import Lock
from urllib.error import HTTPError
from urllib.request import urlopen, Request

import pytz
from PyQt6.QtCore import QThread, pyqtSignal, QMutex

from bs4 import BeautifulSoup
import datetime

from openpyxl.reader.excel import load_workbook

from main import read_excel_template, Flat
# from parser import tz


def get_count(thread):
    # print("\tget count")
    url = 'https://www.olx.uz/nedvizhimost/kvartiry'
    # Request(url).add_header('Accept-Encoding', 'identity')
    while True:
        try:
            page = urlopen(url)
            break
        except Exception as arr:
            print(arr, url, "count")
            time.sleep(10)
            continue
    html = page.read().decode('utf-8')
    soup = BeautifulSoup(html, "html.parser")
    curs = (soup.find(name="ul", attrs={"data-testid": "category-count-links"})
            .find_all(name="li")[1].find(name='span'))
    # print(curs)
    # 104 065
    count_str_arr = re.split(r"\xa0", curs.get_text())
    # print(count_str_arr)
    return int(count_str_arr[0]) * 1000 + int(count_str_arr[1])


# print(get_count())

def get_rate(id_thread):
    # print("\tget rate")
    url = 'https://ofb.uz/uz/'
    # Request(url).add_header('Accept-Encoding', 'identity')
    while True:
        try:
            page = urlopen(url)
            break
        except Exception as arr:
            print(arr, url, "rate")
            time.sleep(10)
            continue
    html = page.read().decode('utf-8')
    soup = BeautifulSoup(html, "html.parser")
    curs = soup.find_all(name="div", attrs={"class": "currency"})
    return float(re.search(r"(\d+)\.(\d+)", curs[0].get_text())[0])


def get_all_flats_from_html(url, page, progress, max_page, prev_count, id_thread):  # UZS -сумм., UYE - y.e.
    list_of_flats = []
    # print("\tget_all_flats_from_html", page)
    url = url + f'?currency=UYE&page={page}'
    # req = Request(url)
    # req.add_header('Accept-Encoding', 'identity')
    while True:
        try:
            page = urlopen(url)
            break
        except Exception as arr:
            print(arr, url, "list", id_thread)
            time.sleep(1)
            continue
    html = page.read().decode('utf-8')
    soup = BeautifulSoup(html, "html.parser")
    ads = soup.find_all(name="div", attrs={"data-cy": "l-card"})

    # rate = requests.get("https://cbu.uz/ru/arkhiv-kursov-valyut/json/USD/").json()[0]['Rate']
    rate = get_rate(id_thread)
    for ad in ads:
        # print(math.floor((ads.index(ad)+prev_count) * 100 / len(ads) / max_page))
        progress.emit(math.ceil((ads.index(ad) + prev_count) * 100 / len(ads) / max_page))
        address_with_modified = ad.find(name='p', attrs={"data-testid": "location-date"}).get_text().split(" - ")
        price_uye = ad.find(name='p', attrs={"data-testid": "ad-price"}).get_text().split(" ")
        if price_uye is None:
            price_uye = 1

        square = ad.find(name='div', attrs={"color": "text-global-secondary"}).get_text()
        final_price = 0
        for i in range(0, len(price_uye) - 1):
            final_price += (int(price_uye[len(price_uye) - 2 - i]) * math.pow(1000, i))

        if "Сегодня" in address_with_modified[1]:
            now = datetime.datetime.now()
            address_with_modified[1] = f'{now.day} {now.strftime("%B").lower()} {now.year}г.'
        details = get_details_of_flat("https://www.olx.uz" + ad.a.get("href"), id_thread)
        # print(details['floor'], "floor", ad.a.get("href"))

        flat = Flat(
            price_uye=final_price,
            price_uzs=float(rate) * float(final_price),
            square=square,
            address=address_with_modified[0],
            modified=address_with_modified[1],
            url="https://www.olx.uz" + ad.a.get("href"),
            room=details['room'],
            floor=details['floor'],
            total_floor=details['total_floor'],
            repair=details['repair'],
            is_new_building=details['is_new_building'],
            description=details['description'],
            id=details['id']
        )
        list_of_flats.append(flat)
    return list_of_flats


def get_details_of_flat(url, id_thread):
    # print("\t\tget_deteil", url)
    req = Request(url)
    req.add_header('Accept-Encoding', 'identity')
    while True:
        try:
            page = urlopen(url)
            break
        # except
        except HTTPError as e:
            print(e, url, "details", id_thread)
            if e.code == 404:
                return
            time.sleep(random.randint(1, 10))
            continue
    html = page.read().decode('utf-8')
    soup = BeautifulSoup(html, "html.parser")
    ad = soup.find(name="div", attrs={"data-testid": "main"}).find(name="ul").get_text()  # li p parse
    ad_desc = soup.find(name="div", attrs={"data-cy": "ad_description", }).find(name="div").get_text()
    # print(ad_desc, url)
    id_ = soup.find(name='div', attrs={"data-cy": "ad-footer-bar-section"}).find(name='span').get_text()
    details = {
        "room": re.search(r"комнат: (\d+)", ad),
        "floor": re.search(r"Этаж: (\d+)", ad),
        "total_floor": re.search(r"Этажность дома: (\d+)", ad),
        "repair": re.search(r"Ремонт: ([А-Я][а-я]+\s*[а-я]*)", ad),
        "is_new_building": re.search(r"жилья: ([А-Я][а-я]+)", ad),
        "description": ad_desc.replace("\n", ";"),
        "id": re.search(r"ID: (\d+)", id_)
    }
    # print(re.search(r"ID: (\d+)", id_)[1], url)
    for detail in details:
        if details[detail] is not None:
            if detail != "description":
                details[detail] = details[detail][1]
        else:
            details[detail] = ''
    return details


class OlxParser(QThread):
    updated = pyqtSignal(int)
    throw_exception = pyqtSignal(str)
    throw_info = pyqtSignal(str)
    label = pyqtSignal(str)
    block_export = pyqtSignal(bool, str)
    block_closing = pyqtSignal(bool)
    date = pyqtSignal()
    lock = QMutex()

    def __init__(self, id, result_from_file, path='_internal/output/internal/'):
        super().__init__()
        self.path = path
        self.id = id
        self.results_from_file = result_from_file

    def run(self):
        print(self.id, "started")
        time.sleep(random.randint(1, 10))
        total_count = get_count(self.id)  # todo get from olx
        self.label.emit("Процесс: Обновление OLX")
        self.updated.emit(0)
        self.block_closing.emit(True)
        # if os.path.exists(self.path + "olx.xlsm"):
        #     results_from_file = get_arr_from_excel(self.path + "olx.xlsm")
        # else:
        #     results_from_file = []

        if not os.path.exists(self.path):
            self.throw_exception.emit("Повреждена файловая система! Перезагрузите приложение")

        # header_sheet(sheet)
        url = "https://www.olx.uz/nedvizhimost/kvartiry/prodazha/"
        # req = Request(url)
        # req.add_header('Accept-Encoding', 'identity')
        while True:
            try:
                html = urlopen(url).read().decode('utf-8')
                self.label.emit("Процесс: Обновление OLX")
                break
            except Exception as err:
                # self.throw_info.emit("Проблемы с подключением к сети")
                print(err, url, "start page", self.id)
                self.label.emit("Процесс: Обновление OLX - Переподключение")
                # self.updated.emit(0)
                time.sleep(random.randint(1, 10))
        print(datetime.datetime.now(tz=pytz.timezone('Europe/Moscow')))
        soup = BeautifulSoup(html, "html.parser")
        max_page = soup.find_all(name="li", attrs={"data-testid": "pagination-list-item"})
        max_page = int(max_page[len(max_page) - 1].get_text())
        page = 1
        start = time.time()
        max_page = 1  # TODO test_data
        prev_count = 0
        # print("olx")
        while page <= max_page:

            try:
                results = get_all_flats_from_html(url, page, self.updated, max_page, prev_count, self.id)
                self.label.emit("Процесс: Обновление OLX")
                # break
            except Exception as err:
                self.throw_info.emit("Проблемы с подключением к сети")
                print(err, url, "uncatch")
                self.label.emit("Процесс: Обновление OLX - Переподключение")
                # self.updated.emit(0)
                time.sleep(random.randint(1, 10))
                continue

            results = [one for one in results if one not in self.results_from_file]
            print([one.id for one in results])
            self.lock.lock()
            book = read_excel_template(self.throw_exception)
            sheet = book[book.sheetnames[0]]
            sheet.title = f"{datetime.datetime.now().strftime('%d.%m.%y_%H.%M')}"
            for i in range(0, len(results)):
                sheet.append(results[i].prepare_to_list())
            # time.sleep(0.001)
            book.save(self.path)
            self.lock.unlock()
            print("olx", time.time() - start, page, " page done", self.id)
            page += 1
            prev_count += len(results)
        self.path += f'olx.xlsm'
        self.block_export.emit(True, "olx")
        # if os.path.exists(self.path):
        #     os.remove(self.path)

        self.block_export.emit(False, "olx")
        self.block_closing.emit(False)
        self.updated.emit(100)
        self.label.emit("Процесс: Обновление OLX - Завершено")
        print(self.id, "finished")


# lock = QMutex()
def get_arr_from_excel(name):
    # lock.lock()
    input_book = load_workbook(name)
    ws_input_book = input_book[input_book.sheetnames[0]]
    flats = []
    for row in ws_input_book.iter_rows(min_row=6, max_col=13):
        flats.append(Flat(
            price_uye=float(row[8].value),
            price_uzs=float(row[10].value),
            square=float(row[1].value.__str__().replace(" ", '')),
            address=row[3].value,
            repair=row[4].value,
            is_new_building=row[5].value,
            room=row[6].value,
            url=row[0].value,
            modified=row[7].value,
            floor=row[2].value.split("/")[0],
            total_floor=row[2].value.split("/")[1],
            description=row[12].value
        ))

    # lock.unlock()
    return flats


def filtration(filters, resource):
    results = get_arr_from_excel(resource)
    print(f"filters: {filters}")
    # print(len(results))
    if 'price_min' in filters:
        if 'uzs' in filters:
            results = [result for result in results if result.price_uzs >= filters['price_min']]
        if 'uye' in filters:
            results = [result for result in results if result.price_uye >= filters['price_min']]
    # print(len(results))
    if 'price_max' in filters:
        if 'uzs' in filters:
            results = [result for result in results if result.price_uzs <= filters['price_max']]
        if 'uye' in filters:
            results = [result for result in results if result.price_uye <= filters['price_max']]
    # print(len(results))
    if 'is_new_building' in filters and filters['is_new_building'] != "Не выбрано":
        results = [result for result in results if result.is_new_building == filters['is_new_building']]
    # print(len(results))
    if 'repair' in filters and filters['repair'] != "Не выбрано":
        results = [result for result in results if result.repair == filters['repair']]
    # print(len(results))
    if 'room' in filters and filters['room'] != "Не выбрано":
        results = [result for result in results if result.room == filters['room']]
    # print(len(results))
    if 'square_min' in filters:
        results = [result for result in results if result.square >= filters['square_min']]
    # print(len(results))
    if 'square_max' in filters:
        results = [result for result in results if result.square <= filters['square_max']]
    # print(len(results))
    if 'floor_min' in filters:
        results = [result for result in results if result.floor >= filters['floor_min']]
    # print(len(results))
    if 'floor_max' in filters:
        results = [result for result in results if int(result.floor) <= int(filters['floor_max'])]
    # print(len(results))
    if 'total_floor_min' in filters:
        results = [result for result in results if result.total_floor >= filters['total_floor_min']]
    # print(len(results))
    if 'total_floor_max' in filters:
        results = [result for result in results if int(result.total_floor) <= int(filters['total_floor_max'])]
    # print(len(results))
    if 'keywords' in filters:
        old = results
        results = []
        for result in old:
            for keyword in filters['keywords']:
                if keyword in (result.description.lower() + result.address.lower()) and not (result in results):
                    results.append(result)
                    # print(result.url)
    return results


def fill_filtered_data(sheet, results, throw_info, name):
    if len(results) == 0:
        throw_info.emit(f"Не найдены данные по запросу для {name}")
        return
    for i in range(0, len(results)):
        sheet.append(results[i].prepare_to_list())
