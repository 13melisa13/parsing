from PyQt6.QtWidgets import QMessageBox
from openpyxl import load_workbook

from olx_parsing import Flat


def get_arr_from_excel(name):
    input_book = load_workbook(name)
    ws_input_book = input_book[input_book.sheetnames[0]]
    flats = []
    for row in ws_input_book.iter_rows(min_row=6, max_col=13):
        flats.append(Flat(
            price_uye=float(row[8].value),
            price_uzs=float(row[10].value),
            square=float(row[1].value.__str__().replace(" ", '')),
            address=row[3].value,
            repair=row[4].value,
            is_new_building=row[5].value,
            room=row[6].value,
            url=row[0].value,
            modified=row[7].value,
            floor=row[2].value.split("/")[0],
            total_floor=row[2].value.split("/")[1],
            description=row[12].value
        ))
    return flats


def filtration(filters, resource):
    results = get_arr_from_excel(resource)
    print(f"filters: {filters}")
    # print(len(results))
    if 'price_min' in filters:
        if 'uzs' in filters:
            results = [result for result in results if result.price_uzs >= filters['price_min']]
        if 'uye' in filters:
            results = [result for result in results if result.price_uye >= filters['price_min']]
    # print(len(results))
    if 'price_max' in filters:
        if 'uzs' in filters:
            results = [result for result in results if result.price_uzs <= filters['price_max']]
        if 'uye' in filters:
            results = [result for result in results if result.price_uye <= filters['price_max']]
    # print(len(results))
    if 'is_new_building' in filters and filters['is_new_building'] != "Не выбрано":
        results = [result for result in results if result.is_new_building == filters['is_new_building']]
    # print(len(results))
    if 'repair' in filters and filters['repair'] != "Не выбрано":
        results = [result for result in results if result.repair == filters['repair']]
    # print(len(results))
    if 'room' in filters and filters['room'] != "Не выбрано":
        results = [result for result in results if result.room == filters['room']]
    # print(len(results))
    if 'square_min' in filters:
        results = [result for result in results if result.square >= filters['square_min']]
    # print(len(results))
    if 'square_max' in filters:
        results = [result for result in results if result.square <= filters['square_max']]
    # print(len(results))
    if 'floor_min' in filters:
        results = [result for result in results if result.floor >= filters['floor_min']]
    # print(len(results))
    if 'floor_max' in filters:
        results = [result for result in results if int(result.floor) <= int(filters['floor_max'])]
    # print(len(results))
    if 'total_floor_min' in filters:
        results = [result for result in results if result.total_floor >= filters['total_floor_min']]
    # print(len(results))
    if 'total_floor_max' in filters:
        results = [result for result in results if int(result.total_floor) <= int(filters['total_floor_max'])]
    # print(len(results))
    if 'keywords' in filters:
        old = results
        results = []
        for result in old:
            for keyword in filters['keywords']:
                if keyword in (result.description.lower() + result.address.lower()) and not (result in results):
                    results.append(result)
                    # print(result.url)
    return results


def fill_filtered_data(sheet, results, throw_info, name):
    if len(results) == 0:
        throw_info.emit(f"Не найдены данные по запросу для {name}")
        return
    for i in range(0, len(results)):
        sheet.append(results[i].prepare_to_list())
        # sheet.append([results[i].price_uye,
        #               results[i].price_per_meter_uye,
        #               results[i].price_uzs,
        #               results[i].price_per_meter_uzs,
        #               results[i].square,
        #               f'{results[i].floor}/{results[i].total_floor}',
        #               results[i].address,
        #               results[i].repair,
        #               results[i].is_new_building,
        #               results[i].room,
        #               results[i].url,
        #               results[i].modified])
